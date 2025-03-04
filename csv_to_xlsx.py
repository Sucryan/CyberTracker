import sys
import csv
import openpyxl
import datetime

def parse_datetime_isoformat(date_str):
    """
    嘗試解析日期/時間字串，支援含 'T' 的 ISO 格式。
    例如：2023-06-29T12:34:56 或 2025-02-15T02:42:55.637
    若尾端有 'Z' 則先去除，再用 fromisoformat。
    傳回 datetime.datetime 物件；若解析失敗傳回 None。
    """
    date_str = date_str.strip()
    if date_str.endswith("Z"):
        date_str = date_str[:-1]
    try:
        return datetime.datetime.fromisoformat(date_str)
    except ValueError:
        return None

def parse_date_yyyy_mm_dd(date_str):
    """
    嘗試解析 'YYYY-MM-DD' 格式 (不含時間)，
    傳回 datetime.date 物件；若失敗傳回 None。
    """
    date_str = date_str.strip()
    try:
        dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
        return dt.date()
    except ValueError:
        return None

def parse_any_date(date_str):
    """
    通用的解析函式，如果同時可能出現
    2025-02-15T02:42:55 或 2025-02-21 (不含 'T')
    就可用此函式先嘗試解析成 datetime，否則再解析成 date。
    """
    if not date_str.strip():
        return None
    dt = parse_datetime_isoformat(date_str)
    if dt:
        return dt
    d = parse_date_yyyy_mm_dd(date_str)
    if d:
        return d
    return None

def convert_csv_to_xlsx(input_csv, output_xlsx):
    """
    針對 CSV 三個欄位，做指定的格式：
      1. 詐騙網站創建日期 → yyyy-mm-ddThh:mm:ss (以 datetime 存)
      2. 接獲通報日期、停止解析日期 → yyyymmdd (以 date 存)
    其餘欄位保持文字。
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # 讀取 CSV 所有行
    with open(input_csv, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        print("[WARNING] CSV 為空")
        return

    # 先寫表頭
    header = rows[0]
    ws.append(header)

    # 尋找三個目標欄位的索引
    col_index_creation = None  # "詐騙網站創建日期"
    col_index_report   = None  # "接獲通報日期"
    col_index_stop     = None  # "停止解析日期"

    for i, col_name in enumerate(header):
        if col_name.strip() == "詐騙網站創建日期":
            col_index_creation = i
        elif col_name.strip() == "接獲通報日期":
            col_index_report = i
        elif col_name.strip() == "停止解析日期":
            col_index_stop = i

    # 寫入資料，並針對三個欄位做相應解析
    for row in rows[1:]:
        new_row = row.copy()
        # 1) 詐騙網站創建日期 → 解析成 datetime，格式 yyyy-mm-ddThh:mm:ss
        if col_index_creation is not None and len(row) > col_index_creation:
            dt = parse_datetime_isoformat(row[col_index_creation])
            if dt:
                new_row[col_index_creation] = dt  # datetime 物件

        # 2) 接獲通報日期、3) 停止解析日期 → 解析成 date，格式 yyyymmdd
        #   但有可能是 'YYYY-MM-DD' 或 'YYYY-MM-DDTHH:mm:ss'
        if col_index_report is not None and len(row) > col_index_report:
            dd = parse_any_date(row[col_index_report])
            if dd:
                # 若 parse_any_date 傳回 datetime，就取其 date()
                if isinstance(dd, datetime.datetime):
                    new_row[col_index_report] = dd.date()
                else:
                    new_row[col_index_report] = dd

        if col_index_stop is not None and len(row) > col_index_stop:
            dd = parse_any_date(row[col_index_stop])
            if dd:
                if isinstance(dd, datetime.datetime):
                    new_row[col_index_stop] = dd.date()
                else:
                    new_row[col_index_stop] = dd

        ws.append(new_row)

    # 寫好資料後，設定 number_format
    #   - 詐騙網站創建日期 => yyyy-mm-ddThh:mm:ss
    #   - 接獲通報日期、停止解析日期 => yyyymmdd
    max_row = ws.max_row

    def set_col_format(col_idx, row_start, date_format):
        """把 worksheet 的第 col_idx+1 欄 (1-based) 從 row_start 到 max_row 設為 date_format"""
        if col_idx is None:
            return
        for r in range(row_start, max_row + 1):
            cell = ws.cell(row=r, column=col_idx + 1)
            val = cell.value
            if isinstance(val, (datetime.datetime, datetime.date)):
                cell.number_format = date_format

    # 1) 創建日期 => "yyyy-mm-ddThh:mm:ss"
    set_col_format(col_index_creation, 2, 'yyyy-mm-dd"T"hh:mm:ss')

    # 2) 接獲通報日期 => "yyyymmdd"
    set_col_format(col_index_report, 2, 'yyyymmdd')

    # 3) 停止解析日期 => "yyyymmdd"
    set_col_format(col_index_stop, 2, 'yyyymmdd')

    wb.save(output_xlsx)
    print(f"[INFO] 已輸出 XLSX：{output_xlsx}")

# ================
# 主程式
# ================
if __name__=="__main__":
    if len(sys.argv) < 3:
        print("用法：python csv_to_xlsx.py <input.csv> <output.xlsx>")
        sys.exit(1)
    input_csv = sys.argv[1]
    output_xlsx = sys.argv[2]
    convert_csv_to_xlsx(input_csv, output_xlsx)
